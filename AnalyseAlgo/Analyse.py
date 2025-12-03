import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
from openpyxl import load_workbook

def run_analysis():
    # Выбор Excel файла
    file_path = filedialog.askopenfilename(
        title="Выберите Excel файл с ордерами",
        filetypes=[("Excel files", "*.xlsx *.xls")]
    )
    if not file_path:
        return

    try:
        wb = load_workbook(filename=file_path, data_only=True)
        sheet = wb.active  # Берём первый лист
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)

        df = pd.DataFrame(data[1:], columns=data[0])
        df.columns = df.columns.str.strip()

        required_cols = ['Realized PNL', 'Fee', 'Pair', 'Type', 'AvgPrice', 'Time(UTC+8)']
        missing = [c for c in required_cols if c not in df.columns]
        if missing:
            messagebox.showerror("Ошибка",
                                 f"В файле отсутствуют обязательные поля:\n{', '.join(missing)}")
            return

        numeric_cols = ['Realized PNL', 'Fee', 'AvgPrice']
        for col in numeric_cols:
            df[col] = pd.to_numeric(df[col], errors='coerce')  

        # Фильтруем Close Long
        df_open_long = df[df['Type'] == 'Close Long'].copy()

        df_open_long['Time(UTC+8)'] = pd.to_datetime(df_open_long['Time(UTC+8)'], errors='coerce')
        df_open_long['TradeDate'] = df_open_long['Time(UTC+8)'].dt.date
        df_open_long['YearMonth'] = df_open_long['Time(UTC+8)'].dt.to_period('M')

        # Метрики
        profit_col = 'Realized PNL'
        fee_col = 'Fee'
        pair_col = 'Pair'
        avg_price_col = 'AvgPrice'

        pnl_by_pair = df_open_long.groupby(pair_col)[profit_col].sum().sort_values(ascending=False)
        pnl_by_day = df_open_long.groupby('TradeDate')[profit_col].sum()
        pnl_by_month = df_open_long.groupby('YearMonth')[profit_col].sum()

        total_fee = df_open_long[fee_col].sum()
        average_price = df_open_long[avg_price_col].mean()
        total_orders = len(df_open_long)
        max_pnl = df_open_long[profit_col].max()
        min_pnl = df_open_long[profit_col].min()
        profitable_trades = (df_open_long[profit_col] > 0).sum()
        losing_trades = (df_open_long[profit_col] < 0).sum()
        zero_trades = (df_open_long[profit_col] == 0).sum()
        percent_profitable = profitable_trades / total_orders * 100 if total_orders > 0 else 0
        average_pnl = df_open_long[profit_col].mean()

        # Формируем результат
        result = f"""
*** РЕЗУЛЬТАТ АНАЛИЗА ***

Средняя прибыль/убыток: {average_pnl}
Процент прибыльных сделок: {percent_profitable:.2f}%
Прибыльные сделки: {profitable_trades}
Убыточные сделки: {losing_trades}
Нулевые сделки: {zero_trades}

Максимальная прибыль: {max_pnl}
Максимальный убыток: {min_pnl}

Общая комиссия: {total_fee}
Средняя цена: {average_price}
Количество ордеров: {total_orders}

=== PNL по монетам ===
{pnl_by_pair}

=== PNL по дням ===
{pnl_by_day}

=== PNL по месяцам ===
{pnl_by_month}
"""

        text_box.delete(1.0, tk.END)
        text_box.insert(tk.END, result)

    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось прочитать файл Excel:\n{e}")


# === GUI ===
window = tk.Tk()
window.title("Crypto Order Analyzer")
window.geometry("900x700")

btn = tk.Button(window, text="Выбрать файл и запустить анализ", command=run_analysis, font=("Arial", 14))
btn.pack(pady=10)

text_box = scrolledtext.ScrolledText(window, width=110, height=35, font=("Consolas", 10))
text_box.pack()

window.mainloop()

